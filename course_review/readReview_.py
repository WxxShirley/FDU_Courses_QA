import csv 
from openpyxl import load_workbook
import jieba
from wordcloud import WordCloud

import nltk
from nltk import *
nltk.download('punkt')


"""
workbook = load_workbook('review.xlsx')
booksheet = workbook.active

rows = booksheet.rows

i = 0
courses = []
for row in rows :
   i = i + 1
   course = []
   course.append(booksheet.cell(row=i,column=1).value)
   course.append(booksheet.cell(row=i,column=2).value)
   courses.append(course)
  
with open('review.csv','w') as csvfile:
    writer = csv.writer(csvfile)

    #writer.writerow(["cid","title","rating","school"])
    writer.writerows(courses)
"""

jieba.add_word("给分",tag = "v")
jieba.add_word("给分不好",tag = 'v')
jieba.add_word("不点名",tag = 'v')
jieba.add_word("无限b+",tag = 'v')
jieba.add_word('拿了A',tag = 'v')


neg_content , pos_content = "",""

csv_file = open("review.csv")
csv_reader_lines = csv.reader(csv_file)
i = 0
pos_count ,neg_count = 0,0

for line in csv_reader_lines:
   txt, label = line[0] , line[1]
   if line[1] == '0':
      words = jieba.cut(txt)
      neg_content += ' '.join(jieba.cut(txt))
      #neg_content += txt
      neg_count += 1
   if line[1] == '1':
      pos_content += ' '.join(jieba.cut(txt))
      pos_count += 1

#file_write_obj = open("pos.txt",'w')
#file_write_obj.write(pos_content)

#file_write_obj = open("neg.txt",'w')
#file_write_obj.write(neg_content)


# Word Cloud Analysis


font_ = "SimHei.ttf"

wordCloud = WordCloud(font_path = font_,background_color='white',max_font_size=60).generate(pos_content)
img = wordCloud.to_image()
img.show()
img.save("positive_reviews.jpg")

wordCloud = WordCloud(font_path = font_,background_color='white',max_font_size=60).generate(neg_content)
img = wordCloud.to_image()
img.show()
img.save("negative_reviews.jpg")


#print(pos_count,neg_count)


