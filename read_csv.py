import csv 
from openpyxl import load_workbook


workbook = load_workbook('Course.xlsx')
booksheet = workbook.active

rows = booksheet.rows

i = 0
courses = []
for row in rows :
   i = i + 1
   course = []
   course.append(booksheet.cell(row=i,column=1).value)
   course.append(booksheet.cell(row=i,column=2).value)
   course.append(booksheet.cell(row=i,column=3).value)
   course.append(booksheet.cell(row=i,column=4).value)
   courses.append(course)

with open('Course.csv','w') as csvfile:
    writer = csv.writer(csvfile)

    #writer.writerow(["cid","title","rating","school"])
    writer.writerows(courses)



workbook = load_workbook('Course_Type.xlsx')
booksheet = workbook.active

rows = booksheet.rows

i = 0
course_types = []
for row in rows :
   i = i + 1
   type_ = []
   type_.append(booksheet.cell(row=i,column=1).value)
   type_.append(booksheet.cell(row=i,column=2).value)
   course_types.append(type_)

with open('Course_Type.csv','w') as csvfile:
    writer = csv.writer(csvfile)

    #writer.writerow(["cid","title","rating","school"])
    writer.writerows(course_types)



workbook = load_workbook('Course2Type.xlsx')
booksheet = workbook.active

rows = booksheet.rows

i = 0
course_types = []
for row in rows :
   i = i + 1
   type_ = []
   type_.append(booksheet.cell(row=i,column=1).value)
   type_.append(booksheet.cell(row=i,column=2).value)
   course_types.append(type_)

with open('Course2Type.csv','w') as csvfile:
    writer = csv.writer(csvfile)

    writer.writerows(course_types)



workbook = load_workbook('Teacher.xlsx')
booksheet = workbook.active

rows = booksheet.rows

i = 0
teachers = []
for row in rows :
   i = i + 1
   teacher = []
   teacher.append(booksheet.cell(row=i,column=1).value)
   teacher.append(booksheet.cell(row=i,column=2).value)
   teachers.append(teacher)

with open('Teacher.csv','w') as csvfile:
    writer = csv.writer(csvfile)

    writer.writerows(teachers)



workbook = load_workbook('Teacher2Course.xlsx')
booksheet = workbook.active

rows = booksheet.rows

i = 0

t2cs = []
for row in rows:
    i=i+1
    t2c = []
    t2c.append(booksheet.cell(row=i,column=1).value)
    t2c.append(booksheet.cell(row=i,column=2).value)
    t2cs.append(t2c)

with open('Teacher2Course.csv','w') as csvfile:
    writer = csv.writer(csvfile)

    writer.writerows(t2cs)




workbook = load_workbook('School.xlsx')
booksheet = workbook.active

rows = booksheet.rows

i = 0

schools = []
for row in rows:
    i = i+1
    school = []
    school.append(booksheet.cell(row=i,column=1).value)
    school.append(booksheet.cell(row=i,column=2).value)
    schools.append(school)

with open('School.csv','w') as csvfile:
    writer = csv.writer(csvfile)
    writer.writerows(schools)




workbook = load_workbook('School2Course.xlsx')
booksheet = workbook.active

rows = booksheet.rows

i = 0

s2cs = []
for row in rows:
    i=i+1
    s2c = []
    s2c.append(booksheet.cell(row=i,column=1).value)
    s2c.append(booksheet.cell(row=i,column=2).value)
    s2cs.append(s2c)

with open('School2Course.csv','w') as csvfile:
    writer = csv.writer(csvfile)

    writer.writerows(s2cs)


workbook = load_workbook('Teacher2School.xlsx')
booksheet = workbook.active

rows = booksheet.rows

i = 0

t2ss = []
for row in rows:
    i=i+1
    t2s = []
    t2s.append(booksheet.cell(row=i,column=1).value)
    t2s.append(booksheet.cell(row=i,column=2).value)
    t2ss.append(t2s)

with open('Teacher2School.csv','w') as csvfile:
    writer = csv.writer(csvfile)

    writer.writerows(t2ss)